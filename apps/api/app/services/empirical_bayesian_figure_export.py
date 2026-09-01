from __future__ import annotations

from typing import Any

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference


def append_dsem_plot_sheets(
    workbook: Workbook,
    parameters: list[dict[str, Any]],
) -> None:
    if not parameters:
        return
    data_sheet = workbook.create_sheet("Bayesian-DSEM绘图数据")
    data_sheet.sheet_view.showGridLines = False
    chart_sheet = workbook.create_sheet("Bayesian-DSEM附录图")
    chart_sheet.sheet_view.showGridLines = False
    for parameter_index, parameter in enumerate(parameters):
        chains = parameter.get("chains", [])
        if not chains:
            continue
        start_column = 1 + parameter_index * (len(chains) + 4)
        data_sheet.cell(1, start_column, parameter.get("label"))
        data_sheet.cell(2, start_column, "迭代")
        for chain_index, chain in enumerate(chains, start=1):
            data_sheet.cell(2, start_column + chain_index, f"Chain {chain.get('chain')}")
        iterations = chains[0].get("iterations", [])
        for row_index, iteration in enumerate(iterations, start=3):
            data_sheet.cell(row_index, start_column, iteration)
            for chain_index, chain in enumerate(chains, start=1):
                values = chain.get("values", [])
                value_index = row_index - 3
                data_sheet.cell(
                    row_index,
                    start_column + chain_index,
                    values[value_index] if value_index < len(values) else None,
                )
        trace = LineChart()
        trace.title = f"{parameter.get('label')}：MCMC 迹线"
        trace.y_axis.title = "参数值"
        trace.x_axis.title = "迭代"
        trace.height = 7
        trace.width = 13
        trace.add_data(
            Reference(
                data_sheet,
                min_col=start_column + 1,
                max_col=start_column + len(chains),
                min_row=2,
                max_row=2 + len(iterations),
            ),
            titles_from_data=True,
        )
        trace.set_categories(
            Reference(
                data_sheet,
                min_col=start_column,
                min_row=3,
                max_row=2 + len(iterations),
            )
        )
        chart_row = 1 + parameter_index * 15
        chart_sheet.add_chart(trace, f"A{chart_row}")

        values = [
            float(value)
            for chain in chains
            for value in chain.get("values", [])
            if value is not None
        ]
        if not values:
            continue
        bin_count = 36
        minimum = min(values)
        maximum = max(values)
        if maximum <= minimum:
            minimum -= 0.5
            maximum += 0.5
        width = (maximum - minimum) / bin_count
        counts = [0] * bin_count
        for value in values:
            index = min(bin_count - 1, max(0, int((value - minimum) / width)))
            counts[index] += 1
        density_column = start_column
        density_start_row = 5 + len(iterations)
        data_sheet.cell(density_start_row, density_column, "参数值")
        data_sheet.cell(density_start_row, density_column + 1, "后验密度")
        for index, count in enumerate(counts, start=1):
            data_sheet.cell(
                density_start_row + index,
                density_column,
                minimum + (index - 0.5) * width,
            )
            data_sheet.cell(
                density_start_row + index,
                density_column + 1,
                count / (len(values) * width),
            )
        density = LineChart()
        density.title = f"{parameter.get('label')}：后验分布"
        density.y_axis.title = "密度"
        density.x_axis.title = "参数值"
        density.height = 7
        density.width = 13
        density.add_data(
            Reference(
                data_sheet,
                min_col=density_column + 1,
                min_row=density_start_row,
                max_row=density_start_row + bin_count,
            ),
            titles_from_data=True,
        )
        density.set_categories(
            Reference(
                data_sheet,
                min_col=density_column,
                min_row=density_start_row + 1,
                max_row=density_start_row + bin_count,
            )
        )
        chart_sheet.add_chart(density, f"J{chart_row}")
