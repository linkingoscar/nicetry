from __future__ import annotations

import json
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services.tabular_security import escape_spreadsheet_formula


def append_sheet(workbook: Workbook, title: str, rows: list[list[Any]]) -> None:
    sheet = workbook.create_sheet(title=title[:31])
    sheet.sheet_view.showGridLines = False
    for row in rows:
        sheet.append(
            [
                (
                    None
                    if isinstance(value, (dict, list)) and not value
                    else json.dumps(value, ensure_ascii=False)
                    if isinstance(value, (dict, list))
                    else escape_spreadsheet_formula(value)
                )
                for value in row
            ]
        )
    if rows:
        header_border = Border(bottom=Side(style="medium", color="244A3D"))
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="315F4E")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = header_border
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=False)
                if isinstance(cell.value, float):
                    cell.number_format = "0.000"
    for column in range(1, sheet.max_column + 1):
        width = max(
            (len(str(sheet.cell(row, column).value or "")) for row in range(1, sheet.max_row + 1)),
            default=8,
        )
        sheet.column_dimensions[get_column_letter(column)].width = min(max(width + 2, 10), 40)
