from __future__ import annotations

import csv
import hashlib
import json
import logging
import os
import stat
import tempfile
import time
import uuid
from datetime import datetime, timezone
from pathlib import Path, PurePosixPath
from typing import Any, BinaryIO, cast
from zipfile import BadZipFile, ZipFile

import pandas as pd
import pyreadstat
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.services.dataset_import_profile import preview_dataset, profile_variables
from app.services.dataset_repository import DatasetRepository
from app.settings import Settings

logger = logging.getLogger("researchpath")

SUPPORTED_EXTENSIONS = {
    ".csv": "csv",
    ".xlsx": "xlsx",
    ".sav": "sav",
    ".dta": "dta",
    ".por": "por",
}
MAX_UPLOAD_BYTES = 50 * 1024 * 1024
MAX_XLSX_UNCOMPRESSED_BYTES = 250 * 1024 * 1024
MAX_XLSX_ARCHIVE_ENTRIES = 10_000
MAX_XLSX_ROWS = 500_000
MAX_XLSX_COLUMNS = 5_000
MAX_XLSX_CELLS = 5_000_000
MAX_DECODED_DATAFRAME_BYTES = 256 * 1024 * 1024


class DatasetImportError(ValueError):
    pass


def _replace_with_windows_retry(source: Path, destination: Path) -> None:
    """Move a completed file without failing on a short-lived Windows handle."""
    for attempt in range(5):
        try:
            os.replace(source, destination)
            return
        except PermissionError:
            if attempt == 4:
                raise
            time.sleep(0.05 * (2**attempt))


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


_utc_now = utc_now_iso


def _safe_original_name(filename: str) -> str:
    clean = Path(filename).name
    clean = "".join(character for character in clean if character.isprintable()).strip()
    return clean[:240] or "uploaded-data"


def _write_upload_to_temporary(source: BinaryIO, state_root: Path) -> tuple[Path, int, str]:
    temporary_root = state_root / "tmp"
    temporary_root.mkdir(parents=True, exist_ok=True)
    digest = hashlib.sha256()
    size = 0
    temporary_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="wb",
            delete=False,
            dir=temporary_root,
            prefix="upload-",
            suffix=".part",
        ) as destination:
            temporary_path = Path(destination.name)
            while True:
                chunk = source.read(1024 * 1024)
                if not chunk:
                    break
                size += len(chunk)
                if size > MAX_UPLOAD_BYTES:
                    raise DatasetImportError("上传文件超过 50 MB 的 M1 限制")
                digest.update(chunk)
                destination.write(chunk)
    except BaseException:
        if temporary_path is not None:
            temporary_path.unlink(missing_ok=True)
        raise
    assert temporary_path is not None
    if size == 0:
        temporary_path.unlink(missing_ok=True)
        raise DatasetImportError("上传文件为空")
    return temporary_path, size, digest.hexdigest()


def _detect_csv(path: Path) -> tuple[str, str]:
    sample_bytes = path.read_bytes()[:128_000]
    encoding = "utf-8-sig"
    try:
        sample = sample_bytes.decode(encoding)
    except UnicodeDecodeError:
        encoding = "gb18030"
        try:
            sample = sample_bytes.decode(encoding)
        except UnicodeDecodeError as error:
            raise DatasetImportError("CSV 编码无法识别，仅支持 UTF-8 和 GB18030") from error
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=",\t;").delimiter
    except csv.Error:
        delimiter = ","
    return encoding, delimiter


def _validate_tabular_shape(rows: int, columns: int, *, label: str) -> None:
    if rows > MAX_XLSX_ROWS or columns > MAX_XLSX_COLUMNS or rows * columns > MAX_XLSX_CELLS:
        raise DatasetImportError(f"{label} 超过 500000 行、5000 列或 500 万单元格限制")


def _validate_csv_shape(path: Path, encoding: str, delimiter: str) -> None:
    row_count = 0
    column_count = 0
    with path.open("r", encoding=encoding, newline="") as source:
        for row in csv.reader(source, delimiter=delimiter):
            row_count += 1
            column_count = max(column_count, len(row))
            _validate_tabular_shape(max(0, row_count - 1), column_count, label="CSV")


def _validate_decoded_dataframe(dataframe: pd.DataFrame) -> None:
    _validate_tabular_shape(int(dataframe.shape[0]), int(dataframe.shape[1]), label="解码后的数据")
    decoded_bytes = int(dataframe.memory_usage(index=True, deep=True).sum())
    if decoded_bytes > MAX_DECODED_DATAFRAME_BYTES:
        raise DatasetImportError("解码后的数据超过 256 MB 内存限制")


def _unique_columns(columns: list[Any]) -> tuple[list[str], list[str]]:
    seen: dict[str, int] = {}
    unique: list[str] = []
    warnings: list[str] = []
    for index, column in enumerate(columns, start=1):
        base = str(column).strip() or f"unnamed_{index}"
        count = seen.get(base, 0)
        seen[base] = count + 1
        resolved = base if count == 0 else f"{base}__{count + 1}"
        if resolved != base:
            warnings.append(f"重复列名 {base} 已规范化为 {resolved}")
        unique.append(resolved)
    return unique, warnings


def _validate_xlsx_archive(path: Path, selected_sheet: str | None = None) -> None:
    with ZipFile(path) as archive:
        entries = archive.infolist()
        if len(entries) > MAX_XLSX_ARCHIVE_ENTRIES:
            raise DatasetImportError("XLSX 内部文件数量超过安全限制")
        total_uncompressed = 0
        for entry in entries:
            member = PurePosixPath(entry.filename)
            if member.is_absolute() or ".." in member.parts:
                raise DatasetImportError("XLSX 包含不安全的内部路径")
            lowered = entry.filename.lower()
            if lowered.startswith("xl/externallinks/") or lowered.endswith("vbaproject.bin"):
                raise DatasetImportError("XLSX 包含外部链接或宏，已拒绝导入")
            total_uncompressed += entry.file_size
            if total_uncompressed > MAX_XLSX_UNCOMPRESSED_BYTES:
                raise DatasetImportError("XLSX 解压后内容超过 250 MB 安全限制")
            if (
                entry.file_size > 1024 * 1024
                and entry.compress_size > 0
                and entry.file_size / entry.compress_size > 200
            ):
                raise DatasetImportError("XLSX 包含异常高压缩比内容")

    with path.open("rb") as source:
        workbook = load_workbook(
            source,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
        try:
            if not workbook.sheetnames:
                raise DatasetImportError("XLSX 不包含工作表")
            sheet_name = selected_sheet or workbook.sheetnames[0]
            if sheet_name not in workbook.sheetnames:
                raise DatasetImportError(f"工作表 {sheet_name} 不存在于该工作簿中")
            worksheet = workbook[sheet_name]
            rows = int(worksheet.max_row or 0)
            columns = int(worksheet.max_column or 0)
            if (
                rows > MAX_XLSX_ROWS
                or columns > MAX_XLSX_COLUMNS
                or rows * columns > MAX_XLSX_CELLS
            ):
                raise DatasetImportError(
                    f"XLSX 工作表 {sheet_name} 超过 500000 行、5000 列或 500 万单元格限制"
                )
        finally:
            workbook.close()


def _is_qualtrics_csv(path: Path, encoding: str, delimiter: str) -> tuple[bool, list[str]]:
    try:
        with path.open("r", encoding=encoding, newline="") as source:
            reader = csv.reader(source, delimiter=delimiter)
            row0 = next(reader, None)
            row1 = next(reader, None)
            row2 = next(reader, None)
            if row0 and row1 and row2:
                qualtrics_hints = sum(1 for cell in row2 if "ImportId" in cell)
                if qualtrics_hints >= min(3, len(row2)):
                    return True, row1
    except Exception:
        pass
    return False, []


def _read_dataframe(
    path: Path, file_format: str, selected_sheet: str | None = None
) -> tuple[pd.DataFrame, dict[str, Any], list[str]]:
    metadata: dict[str, Any] = {"labels": {}, "valueLabels": {}}
    warnings: list[str] = []
    if file_format == "csv":
        encoding, delimiter = _detect_csv(path)
        _validate_csv_shape(path, encoding, delimiter)
        is_qualtrics, qualtrics_labels = _is_qualtrics_csv(path, encoding, delimiter)
        if is_qualtrics:
            dataframe = pd.read_csv(
                path, sep=delimiter, encoding=encoding, skiprows=[1, 2], low_memory=False
            )
            metadata["labels"] = {
                str(col): str(label)
                for col, label in zip(dataframe.columns, qualtrics_labels, strict=False)
                if label
            }
            warnings.append("自动识别到 Qualtrics CSV 表头，已解析变量标签并跳过元数据行")
        else:
            dataframe = pd.read_csv(path, sep=delimiter, encoding=encoding, low_memory=False)
        metadata.update({"encoding": encoding, "delimiter": delimiter})
    elif file_format == "xlsx":
        _validate_xlsx_archive(path, selected_sheet)
        with pd.ExcelFile(path, engine="openpyxl") as workbook:
            sheet_names = list(workbook.sheet_names)
            if not sheet_names:
                raise DatasetImportError("XLSX 不包含工作表")
            sheet = selected_sheet or sheet_names[0]
            if sheet not in sheet_names:
                raise DatasetImportError(f"工作表 {sheet} 不存在于该工作簿中")
            dataframe = pd.read_excel(
                workbook,
                sheet_name=sheet,
                engine="openpyxl",
            )
        metadata["sheet"] = sheet
        metadata["sheetNames"] = sheet_names
        if len(sheet_names) > 1 and not selected_sheet:
            warnings.append(f"工作簿包含 {len(sheet_names)} 个工作表，已读取第一个：{sheet}")
    elif file_format == "sav":
        _, sav_header = pyreadstat.read_sav(path, metadataonly=True)
        _validate_tabular_shape(
            int(sav_header.number_rows or 0),
            int(sav_header.number_columns or 0),
            label="SAV",
        )
        dataframe, sav_metadata = pyreadstat.read_sav(
            path, apply_value_formats=False, user_missing=True
        )
        metadata["labels"] = {
            str(name): label for name, label in sav_metadata.column_names_to_labels.items() if label
        }
        metadata["valueLabels"] = {
            str(name): {str(key): value for key, value in labels.items()}
            for name, labels in sav_metadata.variable_value_labels.items()
        }
    elif file_format == "dta":
        _, dta_header = pyreadstat.read_dta(path, metadataonly=True)
        _validate_tabular_shape(
            int(dta_header.number_rows or 0),
            int(dta_header.number_columns or 0),
            label="DTA",
        )
        dataframe, dta_metadata = pyreadstat.read_dta(path, apply_value_formats=False)
        metadata["labels"] = {
            str(name): label for name, label in dta_metadata.column_names_to_labels.items() if label
        }
        metadata["valueLabels"] = {
            str(name): {str(key): value for key, value in labels.items()}
            for name, labels in dta_metadata.variable_value_labels.items()
        }
    elif file_format == "por":
        _, por_header = pyreadstat.read_por(path, metadataonly=True)
        _validate_tabular_shape(
            int(por_header.number_rows or 0),
            int(por_header.number_columns or 0),
            label="POR",
        )
        dataframe, por_metadata = pyreadstat.read_por(path, apply_value_formats=False)
        metadata["labels"] = {
            str(name): label for name, label in por_metadata.column_names_to_labels.items() if label
        }
        metadata["valueLabels"] = {
            str(name): {str(key): value for key, value in labels.items()}
            for name, labels in por_metadata.variable_value_labels.items()
        }
    else:
        raise DatasetImportError(f"不支持的文件格式: {file_format}")

    dataframe = cast(pd.DataFrame, dataframe)
    _validate_decoded_dataframe(dataframe)

    if dataframe.shape[1] == 0:
        raise DatasetImportError("数据文件不包含变量列")
    if dataframe.shape[0] == 0:
        raise DatasetImportError("数据文件只有表头，没有数据行")
    unique_columns, column_warnings = _unique_columns(list(dataframe.columns))
    dataframe.columns = unique_columns
    warnings.extend(column_warnings)
    return dataframe, metadata, warnings


_profile_variables = profile_variables
_preview = preview_dataset


def import_dataset(
    source: BinaryIO,
    filename: str,
    settings: Settings,
    repository: DatasetRepository,
    selected_sheet: str | None = None,
) -> dict[str, Any]:
    original_name = _safe_original_name(filename)
    extension = Path(original_name).suffix.lower()
    file_format = SUPPORTED_EXTENSIONS.get(extension)
    if file_format is None:
        raise DatasetImportError("仅支持 CSV、XLSX、SAV、DTA 和 POR 文件")

    temporary_path: Path | None = None
    try:
        temporary_path, size, sha256 = _write_upload_to_temporary(source, settings.state_root)
        dataframe, format_metadata, warnings = _read_dataframe(
            temporary_path, file_format, selected_sheet
        )
        dataset_id = f"dataset_{uuid.uuid4().hex[:16]}"
        dataset_root = settings.state_root / "projects" / "default" / "datasets" / dataset_id
        raw_path = dataset_root / "raw" / f"source{extension}"
        normalized_path = dataset_root / "normalized" / "data.parquet"
        manifest_path = dataset_root / "manifest.json"
        raw_path.parent.mkdir(parents=True, exist_ok=True)
        normalized_path.parent.mkdir(parents=True, exist_ok=True)
        _replace_with_windows_retry(temporary_path, raw_path)
        temporary_path = None

        normalized_temporary = normalized_path.with_suffix(".parquet.tmp")
        dataframe.to_parquet(normalized_temporary, index=False, engine="pyarrow")
        _replace_with_windows_retry(normalized_temporary, normalized_path)
        os.chmod(raw_path, stat.S_IREAD)

        manifest = {
            "schemaVersion": "1.0.0",
            "id": dataset_id,
            "projectId": "default",
            "createdAt": _utc_now(),
            "originalFile": {
                "name": original_name,
                "format": file_format,
                "sizeBytes": size,
                "sha256": sha256,
                **{
                    key: format_metadata[key]
                    for key in ("encoding", "delimiter", "sheet", "sheetNames")
                    if key in format_metadata
                },
            },
            "storage": {
                "raw": raw_path.relative_to(settings.state_root).as_posix(),
                "normalized": normalized_path.relative_to(settings.state_root).as_posix(),
            },
            "rowCount": int(dataframe.shape[0]),
            "columnCount": int(dataframe.shape[1]),
            "variables": _profile_variables(dataframe, format_metadata),
            "preview": _preview(dataframe),
            "warnings": [
                {"code": "IMPORT_NOTE", "severity": "warning", "message": message}
                for message in warnings
            ],
        }
        manifest_path.write_text(
            json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        repository.record_dataset(manifest, manifest_path)
        return repository.get_dataset(dataset_id)
    except (
        DatasetImportError,
        pd.errors.ParserError,
        pyreadstat.ReadstatError,
        BadZipFile,
        InvalidFileException,
        ValueError,
        OSError,
    ) as error:
        if isinstance(error, DatasetImportError):
            raise
        logger.exception("Dataset import failed before the error boundary")
        raise DatasetImportError(
            "数据导入失败，请检查文件内容或格式后重试；诊断信息已写入服务日志。"
        ) from error
    finally:
        if temporary_path is not None:
            try:
                temporary_path.unlink(missing_ok=True)
            except OSError:
                # Cleanup must never replace the actionable import error with a
                # second Windows sharing violation. A later import can safely
                # ignore an orphaned uniquely named temporary file.
                pass
