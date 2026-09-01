"""Single-procedure report presentation, including deliberately omitted sections."""
from __future__ import annotations

from collections.abc import Mapping
from typing import cast

from openpyxl import Workbook

from app.services.empirical_export_support import append_sheet


def _mapping(value: object) -> Mapping[str, object]:
    return cast(Mapping[str, object], value) if isinstance(value, dict) else {}


def _records(value: object) -> list[Mapping[str, object]]:
    return [_mapping(item) for item in value] if isinstance(value, list) else []


def procedure_interpretation(report: Mapping[str, object]) -> tuple[str, str]:
    options = _mapping(report["options"])
    procedure = options["procedure"]
    dependencies = _mapping(report.get("provenance")).get("dependencies")
    text = (
        f"本次仅执行 {procedure}，绑定测量版本 {report['measurementVersionId']}。"
        "各统计表的 N 为实际分析样本量；未选择的方法不构成已执行证据。"
    )
    if isinstance(dependencies, list) and dependencies:
        text += "必要依赖：" + ", ".join(str(item) for item in dependencies) + "。"
    return text, ""


def scope_procedure_workbook(workbook: Workbook, report: Mapping[str, object]) -> None:
    procedure = _mapping(report.get("options")).get("procedure")
    if procedure is None:
        return
    # Sheets from empty compatibility objects must never appear as executed output.
    always_conditional = {
        "描述统计": bool(report.get("descriptives")),
        "频数分布": bool(report.get("frequencies")),
        "方法诊断": procedure in {"efa", "common_method"},
        "信效度": procedure == "validity",
        "EFA载荷": procedure == "efa",
        "测量方法执行": procedure in {"efa", "cfa", "validity"},
    }
    for sheet, keep in always_conditional.items():
        if not keep and sheet in workbook.sheetnames:
            del workbook[sheet]
    if "方法诊断" in workbook.sheetnames:
        sheet = workbook["方法诊断"]
        for index in range(sheet.max_row, 1, -1):
            label = sheet.cell(index, 1).value
            if (procedure == "efa" and label == "Harman第一因子解释率(%)") or (
                procedure == "common_method" and label != "Harman第一因子解释率(%)"
            ):
                sheet.delete_rows(index)
    if "测量方法执行" in workbook.sheetnames:
        sheet = workbook["测量方法执行"]
        selected = {"efa": {"EFA"}, "cfa": {"CFA"}, "validity": {"CFA", "构念效度", "HTMT"}}.get(str(procedure), set())
        for index in range(sheet.max_row, 1, -1):
            if sheet.cell(index, 1).value not in selected:
                sheet.delete_rows(index)
    if procedure != "correlation":
        for sheet in list(workbook.sheetnames):
            if sheet.startswith("相关") or sheet == "论文整合表":
                del workbook[sheet]
    if procedure == "common_method":
        ulmc = _mapping(_mapping(report.get("commonMethodBias")).get("ulmc"))
        rows_ulmc: list[list[object]] = [["模型", "χ²", "df", "CFI", "RMSEA"]]
        if ulmc.get("available"):
            for label, key in (("基准模型", "baselineModel"), ("方法因子模型", "ulmcModel")):
                model = _mapping(ulmc.get(key))
                rows_ulmc.append([label, model.get("chisq"), model.get("df"), model.get("cfi"), model.get("rmsea")])
            comparison = _mapping(ulmc.get("modelComparison"))
            rows_ulmc.extend([["Δχ²", "Δdf", "p", "ΔCFI", "ΔRMSEA"],
                              [comparison.get(key) for key in ("deltaChisq", "deltaDf", "pValue", "deltaCfi", "deltaRmsea")],
                              [ulmc.get("methodologicalWarning")]])
        else:
            rows_ulmc.append(["未完成", ulmc.get("reason")])
        append_sheet(workbook, "ULMC", rows_ulmc)
    if procedure == "reliability":
        rows: list[list[object]] = [["量表", "N", "题项数", "标准化α", "ω", "有序α", "有序ω", "方法", "有序信度不可用原因"]]
        items: list[list[object]] = [["量表", "题项", "校正题总相关", "删除后标准化α", "删除后ω"]]
        for construct in _records(_mapping(report.get("reliability")).get("constructs")):
            stats = _mapping(construct["statistics"])
            rows.append([construct["label"], construct["n"], construct["itemCount"],
                         stats.get("alpha"), stats.get("omega"), stats.get("ordinalAlpha"),
                         stats.get("ordinalOmega"), stats.get("method"), stats.get("ordinalReliabilityReason")])
            items.extend([construct["label"], item["itemId"], item.get("correctedItemTotalCorrelation"),
                          item.get("alphaIfDeleted"), item.get("omegaIfDeleted")]
                         for item in _records(construct.get("items")))
        append_sheet(workbook, "信度分析", rows)
        append_sheet(workbook, "项目诊断", items)
