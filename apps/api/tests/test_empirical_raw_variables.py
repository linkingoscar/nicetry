from __future__ import annotations

from io import BytesIO

import pytest
from _empirical_center_helpers import (
    _await_empirical_job,
    _reference_bh_adjust,
    _reference_holm_adjust,
)
from _sem_calculations_helpers import _ensure_independent_context
from m3_helpers import client
from openpyxl import load_workbook


def raw_dataset(csv: str, types: list[str]) -> dict:
    imported = client.post('/api/v1/datasets/import', files={'file': ('raw.csv', csv.encode(), 'text/csv')})
    assert imported.status_code == 201, imported.text
    dataset = imported.json()
    confirmed = client.put(f"/api/v1/datasets/{dataset['id']}/dictionary", json={
        'variables': [{'id': variable['id'], 'confirmed_type': kind} for variable, kind in zip(dataset['variables'], types, strict=True)]
    })
    assert confirmed.status_code == 200, confirmed.text
    dataset = confirmed.json()
    _ensure_independent_context(dataset)
    return dataset


@pytest.mark.parametrize('procedure,csv,types', [
    ('descriptives', 'age\n20\n25\n30\n35\n40\n', ['continuous']),
    ('frequencies', 'group\nA\nB\nA\nC\nB\n', ['nominal']),
    ('missing', 'age\n20\nNA\n30\n35\n40\n', ['continuous']),
    ('correlation', 'x,y\n1,3\n2,2\n3,7\n4,5\n5,8\n', ['continuous', 'continuous']),
])
def test_raw_procedures_without_measurement_keep_identity_and_export(procedure, csv, types):
    dataset = raw_dataset(csv, types)
    assert client.get(f"/api/v1/datasets/{dataset['id']}/measurement").status_code == 404
    response = client.post(f"/api/v1/datasets/{dataset['id']}/empirical-analysis", json={
        'procedure': procedure, 'analysis_variable_ids': [v['id'] for v in dataset['variables']]
    })
    job = _await_empirical_job(response)
    assert job['measurementVersion'] is None
    assert job['measurementVersionId'] is None
    result = client.get(f"/api/v1/analyses/{job['id']}/result")
    assert result.status_code == 200, result.text
    report = result.json()
    assert report['measurementVersionId'] is None
    assert report['datasetId'] == dataset['id']
    assert report['provenance']['analysisContext']['measurement'] is None
    if procedure == 'descriptives':
        assert report['descriptives'][0]['mean'] == 30
    if procedure == 'frequencies':
        assert report['frequencies']
    base = f"/api/v1/datasets/{dataset['id']}/empirical-analyses/{job['reportId']}"
    assert client.get(base + '/segments/summary').status_code == 200
    exported = client.get(base + '/export')
    assert exported.status_code == 200, exported.text
    workbook = load_workbook(BytesIO(exported.content))
    identity = {row[1]: row[2] for row in workbook["方法与来源"].values if row[0] == "结果身份"}
    assert identity == {"reportId": job["reportId"], "datasetId": dataset["id"], "measurementVersion": None, "measurementVersionId": None}
    assert client.get(f"/api/v1/datasets/{dataset['id']}/measurement").status_code == 404


def test_raw_endpoint_rejects_measurement_methods_before_enqueue():
    dataset = raw_dataset('age\n20\n25\n30\n35\n40\n', ['continuous'])
    response = client.post(f"/api/v1/datasets/{dataset['id']}/empirical-analysis", json={'procedure': 'cfa'})
    assert response.status_code == 422
    assert '测量版本' in response.text


def test_selected_raw_sample_requires_its_own_context_hash():
    dataset = raw_dataset('age\n20\n25\n30\n35\n40\n', ['continuous'])
    dataset_id = dataset['id']
    variable_id = dataset['variables'][0]['id']
    quality = client.post(f'/api/v1/datasets/{dataset_id}/quality-runs', json={'qualityVariableIds': [variable_id], 'durationVariableId': variable_id})
    assert quality.status_code == 201, quality.text
    sample = client.post(f'/api/v1/datasets/{dataset_id}/sample-versions', json={
        'qualityRunId': quality.json()['id'], 'combineOperator': 'or', 'label': '年龄筛选',
        'rules': [{'id': 'rule_age_filter', 'metric': 'duration_seconds', 'operator': 'lt', 'threshold': 30, 'source': 'planned_not_preregistered', 'description': '测试显式样本身份'}],
    })
    assert sample.status_code == 201, sample.text
    base = f'/api/v1/datasets/{dataset_id}/resolved-analysis-context'
    original = client.get(base).json()
    selected = client.get(base, params={'sampleVersionId': sample.json()['id']}).json()
    assert selected['contextHash'] != original['contextHash']
    options = {'procedure': 'descriptives', 'analysis_variable_ids': [variable_id], 'sample_version_id': sample.json()['id'], 'context_hash': original['contextHash']}
    rejected = client.post(f'/api/v1/datasets/{dataset_id}/empirical-analysis', json=options)
    assert rejected.status_code == 409, rejected.text
    assert 'ANALYSIS_CONTEXT_CHANGED' in rejected.text
    options['context_hash'] = selected['contextHash']
    job = _await_empirical_job(client.post(f'/api/v1/datasets/{dataset_id}/empirical-analysis', json=options))
    result = client.get(f"/api/v1/analyses/{job['id']}/result").json()
    assert result['descriptives'][0]['mean'] == 35
    assert result['provenance']['analysisContext']['sample']['id'] == sample.json()['id']


@pytest.mark.parametrize('method', ['holm', 'BH', 'none'])
def test_raw_group_correction_matches_reference_and_excel(method):
    csv = 'group,x,y,z\n' + '\n'.join(f"{i % 2},{i % 7 + (i % 2) * 0.5},{i % 11 + (i % 2) * 0.8},{i % 13}" for i in range(60))
    dataset = raw_dataset(csv, ['binary', 'continuous', 'continuous', 'continuous'])
    job = _await_empirical_job(client.post(f"/api/v1/datasets/{dataset['id']}/empirical-analysis", json={
        'procedure': 'groups', 'analysis_variable_ids': [v['id'] for v in dataset['variables'][1:]],
        'group_variable_id': dataset['variables'][0]['id'], 'group_omnibus_p_adjust': method,
    }))
    report = client.get(f"/api/v1/analyses/{job['id']}/result").json()
    comparison = report['groupComparison']
    raw = [row['pValueRaw'] for row in comparison['results']]
    expected = _reference_holm_adjust(raw) if method == 'holm' else _reference_bh_adjust(raw) if method == 'BH' else raw
    assert comparison['multiplicity']['adjustment'] == method
    assert [row['pValueAdjusted'] for row in comparison['results']] == pytest.approx(expected, abs=1e-12)
    exported = client.get(f"/api/v1/datasets/{dataset['id']}/empirical-analyses/{job['reportId']}/export")
    assert exported.status_code == 200, exported.text
    workbook = load_workbook(BytesIO(exported.content))
    values = [cell.value for sheet in workbook for row in sheet for cell in row]
    for value in expected:
        assert any(isinstance(cell, (int, float)) and abs(cell - value) < 1e-12 for cell in values)
