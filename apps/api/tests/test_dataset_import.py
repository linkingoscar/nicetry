from __future__ import annotations

import hashlib
from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile

import pandas as pd
import pyreadstat
import pytest
from openpyxl import Workbook
from starlette.testclient import TestClient

import app.services.dataset_import as dataset_import
from app.main import app
from app.services.dataset_import import (
    DatasetImportError,
    _profile_variables,
    _read_dataframe,
    _validate_xlsx_archive,
    _write_upload_to_temporary,
)
from app.settings import get_settings

client = TestClient(
    app,
    headers={"X-ResearchPath-Token": app.state.services.settings.session_token},
)


def test_oversized_upload_removes_partial_temporary_file(tmp_path, monkeypatch) -> None:
    monkeypatch.setattr(dataset_import, "MAX_UPLOAD_BYTES", 4)

    with pytest.raises(DatasetImportError, match="50 MB"):
        _write_upload_to_temporary(BytesIO(b"12345"), tmp_path)

    assert list((tmp_path / "tmp").glob("upload-*.part")) == []


def test_csv_shape_is_rejected_before_dataframe_decode(tmp_path, monkeypatch) -> None:
    path = tmp_path / "wide.csv"
    path.write_text("a,b,c\n1,2,3\n4,5,6\n", encoding="utf-8")
    monkeypatch.setattr(dataset_import, "MAX_XLSX_CELLS", 5)

    with pytest.raises(DatasetImportError, match="CSV 超过"):
        _read_dataframe(path, "csv")


def test_csv_profile_infers_core_questionnaire_types(tmp_path) -> None:
    path = tmp_path / "questionnaire.csv"
    path.write_text(
        "respondent_id,condition,satisfaction,score,comment\n"
        "1,A,1,1.2,第一条\n"
        "2,B,2,2.5,第二条\n"
        "3,A,3,3.1,第三条\n"
        "4,B,4,4.8,第四条\n"
        "5,A,5,5.4,第五条\n"
        "6,B,1,6.2,第六条\n"
        "7,A,2,7.3,第七条\n"
        "8,B,3,8.1,第八条\n"
        "9,A,4,9.6,第九条\n"
        "10,B,5,10.4,第十条\n",
        encoding="utf-8",
    )

    dataframe, metadata, warnings = _read_dataframe(path, "csv")
    variables = {
        variable["originalName"]: variable for variable in _profile_variables(dataframe, metadata)
    }

    assert warnings == []
    assert metadata["encoding"] == "utf-8-sig"
    assert variables["respondent_id"]["inferredType"] == "id"
    assert variables["condition"]["inferredType"] == "binary"
    assert variables["satisfaction"]["inferredType"] == "ordinal"
    assert variables["score"]["inferredType"] == "continuous"
    assert variables["comment"]["inferredType"] == "text"


def test_xlsx_reads_first_sheet_and_warns_about_additional_sheets(tmp_path) -> None:
    path = tmp_path / "multi-sheet.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame({"x": [1, 2], "y": [3, 4]}).to_excel(writer, sheet_name="主数据", index=False)
        pd.DataFrame({"notes": ["说明"]}).to_excel(writer, sheet_name="说明", index=False)

    dataframe, metadata, warnings = _read_dataframe(path, "xlsx")

    assert dataframe.shape == (2, 2)
    assert metadata["sheet"] == "主数据"
    assert metadata["sheetNames"] == ["主数据", "说明"]
    assert "已读取第一个" in warnings[0]


def test_api_imports_xlsx_after_reader_releases_windows_file_handle(tmp_path) -> None:
    path = tmp_path / "consumer-questionnaire.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame(
            {
                "respondent_id": range(1, 11),
                "purchase_intention_1": [1, 2, 3, 4, 5, 1, 2, 3, 4, 5],
                "purchase_intention_2": [2, 3, 4, 5, 6, 2, 3, 4, 5, 6],
            }
        ).to_excel(writer, index=False)

    response = client.post(
        "/api/v1/datasets/import",
        files={
            "file": (
                "消费者行为问卷.xlsx",
                BytesIO(path.read_bytes()),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 201, response.text
    dataset = response.json()
    assert dataset["rowCount"] == 10
    assert dataset["columnCount"] == 3
    settings = get_settings()
    assert (settings.state_root / dataset["storage"]["raw"]).read_bytes() == path.read_bytes()


def test_invalid_xlsx_returns_actionable_client_error_instead_of_500() -> None:
    response = client.post(
        "/api/v1/datasets/import",
        files={
            "file": (
                "损坏问卷.xlsx",
                BytesIO(b"not-an-xlsx"),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 422, response.text
    assert "数据导入失败" in response.json()["detail"]["message"]


def test_xlsx_resource_and_external_link_limits_are_enforced(tmp_path) -> None:
    oversized = tmp_path / "oversized-dimension.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.cell(row=500_001, column=1, value=1)
    workbook.save(oversized)
    with pytest.raises(DatasetImportError, match="500000 行"):
        _validate_xlsx_archive(oversized)

    linked = tmp_path / "external-link.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "safe"
    workbook.save(linked)
    with ZipFile(linked, "a", compression=ZIP_DEFLATED) as archive:
        archive.writestr("xl/externalLinks/externalLink1.xml", "<externalLink />")
    with pytest.raises(DatasetImportError, match="外部链接或宏"):
        _validate_xlsx_archive(linked)


def test_sav_preserves_variable_and_value_labels(tmp_path) -> None:
    path = tmp_path / "labels.sav"
    pyreadstat.write_sav(  # type: ignore[arg-type]
        pd.DataFrame({"group": [1, 2, 1], "score": [3.2, 4.1, 2.9]}),
        str(path),
        column_labels={"group": "实验组别", "score": "绩效得分"},
        variable_value_labels={"group": {1: "处理组", 2: "对照组"}},
    )

    dataframe, metadata, warnings = _read_dataframe(path, "sav")

    assert dataframe.shape == (3, 2)
    assert warnings == []
    assert metadata["labels"]["group"] == "实验组别"
    assert metadata["valueLabels"]["group"] == {"1.0": "处理组", "2.0": "对照组"}


def test_api_imports_immutable_version_and_confirms_dictionary() -> None:
    payload = (
        "respondent_id,condition,satisfaction,score\n"
        "1,A,1,1.2\n2,B,2,2.5\n3,A,3,3.1\n4,B,4,4.8\n"
        "5,A,5,5.4\n6,B,1,6.2\n7,A,2,7.3\n8,B,3,8.1\n"
        "9,A,4,9.6\n10,B,5,10.4\n"
    ).encode("utf-8")
    response = client.post(
        "/api/v1/datasets/import",
        files={"file": ("questionnaire.csv", BytesIO(payload), "text/csv")},
    )

    assert response.status_code == 201, response.text
    dataset = response.json()
    assert dataset["rowCount"] == 10
    assert dataset["columnCount"] == 4
    assert dataset["originalFile"]["sha256"] == hashlib.sha256(payload).hexdigest()
    assert dataset["dictionary"]["status"] == "draft"

    settings = get_settings()
    assert (settings.state_root / dataset["storage"]["raw"]).exists()
    assert (settings.state_root / dataset["storage"]["normalized"]).exists()

    updates = [
        {"id": variable["id"], "confirmed_type": variable["inferredType"]}
        for variable in dataset["variables"]
    ]
    confirmed_response = client.put(
        f"/api/v1/datasets/{dataset['id']}/dictionary",
        json={"variables": updates},
    )

    assert confirmed_response.status_code == 200, confirmed_response.text
    confirmed = confirmed_response.json()
    assert confirmed["dictionary"] == {
        "version": 1,
        "confirmedCount": 4,
        "totalCount": 4,
        "status": "confirmed",
    }
    assert all(variable["confirmedType"] for variable in confirmed["variables"])
