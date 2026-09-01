from __future__ import annotations

import json
from io import BytesIO

import numpy as np
import pandas as pd
import pytest
from _empirical_center_helpers import _await_empirical_job
from m3_helpers import _model_dataset, client
from openpyxl import load_workbook

from app.api.dto.empirical_spec_builder import build_empirical_analysis_options
from app.api.schemas import EmpiricalAnalysisRequest
from app.services.empirical_context_gate import empirical_capability_slices
from app.services.empirical_export import empirical_report_path
from app.services.empirical_options_validator import EmpiricalAnalysisError
from app.services.empirical_procedures import validate_procedure
from app.settings import get_settings


def _run(dataset: dict, measurement: dict, procedure: str, **kwargs: object) -> dict:
    response = client.post(
        f"/api/v1/datasets/{dataset['id']}/measurements/{measurement['version']}/empirical-analysis",
        json={"procedure": procedure, **kwargs},
    )
    job = _await_empirical_job(response)
    return json.loads(empirical_report_path(
        dataset["id"], measurement["version"], job["reportId"], get_settings()
    ).read_text(encoding="utf-8"))


def test_single_correlation_scopes_variables_and_does_not_execute_measurement() -> None:
    dataset, measurement = _model_dataset()
    report = _run(dataset, measurement, "correlation", analysis_variable_ids=["scale_x", "scale_y"])
    assert [v["id"] for v in report["correlations"]["variables"]] == ["scale_x", "scale_y"]
    frame = pd.read_parquet(get_settings().state_root / measurement["derivedDataset"]["storage"])
    assert report["correlations"]["coefficients"][0][1] == pytest.approx(
        np.corrcoef(frame["scale_x"], frame["scale_y"])[0, 1], abs=1e-8
    )
    for key in ("efa", "cfa", "validity", "commonMethodBias", "measurementInvariance"):
        assert report[key]["reason"] == "not_requested"
    assert report["descriptives"] == []
    assert report["hierarchicalRegression"] is None
    assert report["provenance"]["htmtBootstrap"]["replicates"] == 0
    assert "CFA_UNAVAILABLE" not in {w["code"] for w in report["warnings"]}
    assert "CFA 未能估计" not in report["academicInterpretation"]
    response = client.get(
        f"/api/v1/datasets/{dataset['id']}/measurements/{measurement['version']}/empirical-analyses/{report['reportId']}/export"
    )
    assert response.status_code == 200, response.text
    workbook = load_workbook(BytesIO(response.content))
    assert not {"信效度", "EFA载荷", "方法诊断", "描述统计", "频数分布"} & set(workbook.sheetnames)
    assert "相关矩阵" in workbook.sheetnames


@pytest.mark.parametrize("procedure", ["descriptives", "frequencies", "missing", "efa", "cfa", "reliability", "groups", "regression", "relative_importance", "response_surface", "common_method", "invariance", "validity"])
def test_single_procedures_keep_unselected_estimators_unrun(procedure: str) -> None:
    dataset, measurement = _model_dataset()
    kwargs: dict[str, object] = {}
    if procedure in {"descriptives", "frequencies", "missing", "groups"}:
        kwargs["analysis_variable_ids"] = ["scale_x"]
    if procedure in {"efa", "cfa", "reliability", "common_method", "invariance", "validity"}:
        kwargs["construct_ids"] = [c["id"] for c in measurement["constructs"][:2]]
    if procedure in {"groups", "invariance"}:
        kwargs["group_variable_id"] = next(v["id"] for v in dataset["variables"] if v["originalName"] == "group")
    if procedure in {"regression", "relative_importance", "response_surface"}:
        kwargs["outcome_variable_id"] = "scale_y"
        kwargs["predictor_variable_ids" if procedure != "response_surface" else "response_surface_predictor_ids"] = ["scale_x", "scale_m"]
    report = _run(dataset, measurement, procedure, **kwargs)
    assert report["options"]["procedure"] == procedure
    assert report["correlations"]["variables"] == []
    if procedure != "efa":
        assert report["efa"]["reason"] == "not_requested"
    if procedure not in {"cfa", "validity"}:
        assert report["cfa"]["reason"] == "not_requested"
    if procedure == "reliability":
        assert len(report["reliability"]["constructs"]) == 2
        assert report["reliability"]["constructs"][0]["statistics"]["alpha"] is not None
    if procedure == "descriptives":
        assert len(report["descriptives"]) == 1
    if procedure == "regression":
        assert report["hierarchicalRegression"]["relativeImportance"] is None
    if procedure == "relative_importance":
        assert report["hierarchicalRegression"]["relativeImportance"]["available"] is True
    if procedure == "validity":
        assert report["provenance"]["dependencies"] == ["cfa"]


def test_scope_validation_rejects_other_methods_and_bad_variables() -> None:
    metadata = {"variables": [{"id": "x", "type": "continuous"}, {"id": "y", "type": "continuous"}], "constructs": []}
    for options in (
        {"procedure": "descriptives", "analysisVariableIds": ["x"], "outcomeVariableId": "y"},
        {"procedure": "correlation", "analysisVariableIds": ["x"]},
        {"procedure": "descriptives", "analysisVariableIds": ["missing"]},
        {"procedure": "correlation", "analysisVariableIds": ["x", "y"], "correlationMethod": "partial"},
        {"procedure": "correlation", "analysisVariableIds": ["x", "y"], "controlVariableIds": ["x"]},
        {"procedure": "groups", "analysisVariableIds": ["x"], "groupVariableId": "x"},
    ):
        with pytest.raises(EmpiricalAnalysisError):
            validate_procedure(metadata, options)
    request = EmpiricalAnalysisRequest(procedure="correlation", analysis_variable_ids=["x", "y"])
    options = build_empirical_analysis_options(request)
    assert empirical_capability_slices(options) == ("empirical.cross_sectional.overview",)
    assert options["analysisVariableIds"] == ["x", "y"]


def test_legacy_requests_still_declare_bundle_capabilities() -> None:
    assert empirical_capability_slices({}) == (
        "empirical.cross_sectional.overview", "empirical.cross_sectional.measurement"
    )


def test_applicable_capabilities_preserves_oracle_independence_contract() -> None:
    dataset, _ = _model_dataset()
    response = client.get(f"/api/v1/datasets/{dataset['id']}/applicable-capabilities")
    assert response.status_code == 200, response.text
    evidence = [row["validationEvidence"] for row in response.json()["capabilities"]]
    assert evidence and all("oracleIndependence" in row for row in evidence)
